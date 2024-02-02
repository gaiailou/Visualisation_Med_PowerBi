"""
Created on Thu Mar  9 08:57:16 2023

@author: nicol
"""
import openpyxl # pour manipuler des fichiers Excel
from habanero import Crossref # pour interagir avec l'API Crossref
from requests.exceptions import HTTPError # pour gérer les erreurs de requête HTTP
import concurrent.futures # pour exécuter plusieurs tâches en parallèle
import threading # pour le verrouillage des ressources partagées

import pandas as pd


def get_author_info(doi):
    cr = Crossref() # initialisation d'un objet Crossref pour interagir avec l'API
    pub = None # initialisation de la variable de résultat
    try:
        pub = cr.works(ids=doi) # obtenir les informations de la publication avec l'ID DOI
        authors = pub['message']['author'] # extraire la liste des auteurs de la réponse JSON
        author_list = []
        for author in authors:
            author_name = author['given'] + ' ' + author['family'] # concaténer le prénom et le nom de famille de chaque auteur
            author_list.append(author_name) # ajouter le nom complet à la liste des noms d'auteurs
        return author_list
    except KeyError: # si la clé demandée n'est pas trouvée dans le dictionnaire JSON et retourner None
        return None
    except HTTPError as http_err: # si une erreur HTTP se produit -> afficher le message d'erreur et retourner None
        print(f"HTTP error occurred: {http_err}")
        return None
    except Exception as err: # si une autre exception se produit -> # afficher le message d'erreur et retourner None 
        print(f"An error occurred: {err}")
        return None


def process_row(row):
    global lock, ws # déclarer les variables globales pour le verrouillage et la feuille de calcul Excel
    doi = ws.cell(row=row, column=6).value # obtenir la valeur de DOI à partir de la colonne 6

    # Vérifier si la valeur de doi est vide
    if doi is None or doi.strip() == '':
        with lock:
            ws.cell(row=row, column=18).value = "null"
        return None

    # Récupérer les noms d'auteurs correspondants
    author_list = get_author_info(doi)

    # Concaténer les noms d'auteurs en une chaîne de caractères séparée par des virgules
    author_str = ', '.join(author_list) if author_list else "null"

    # Écrire les noms d'auteurs dans la colonne 18
    with lock:
        ws.cell(row=row, column=18).value = author_str

    # Afficher les noms d'auteurs ajoutés dans la console
    print(f"Auteurs ajoutés pour la ligne {row}: {author_str}")
    return author_str



#import du fichier excel (ne pas oublier de changer le nom du fichier)
f1 = pd.read_excel('datanoproblem.xlsx', sheet_name='1 - ClinicalTrials_ObsStudies')
f2 = pd.read_excel('datanoproblem.xlsx', sheet_name='2 - ClinicalTrials_RandTrials')
f3 = pd.read_excel('datanoproblem.xlsx', sheet_name='3 - Publications_ObsStudies')
f4 = pd.read_excel('datanoproblem.xlsx', sheet_name='4 - Publications_RandTrials')
print(f1['date'].dtype)     
print(f2['date'].dtype)   

#Ajout de précision de la provenance des données avant concaténation
f1['provenance'] = 'ObsStudies'
f2['provenance'] = 'RandTrials'
f3['provenance'] = 'ObsStudies'
f4['provenance'] = 'RandTrials'
#Partie concaténation
Trials = pd.concat([f1, f2], axis=0)
Publication = pd.concat([f3, f4], axis=0)

#changement des type vers date
Trials['dateInserted'] = Trials['dateInserted'].dt.date
Trials['date'] = Trials['date'].dt.date
Publication['dateInserted'] = Publication['dateInserted'].dt.date
Publication['datePublished']  = Publication['datePublished'].dt.date


#export des fichier
Trials.to_excel('Trials.xlsx', index=False, sheet_name="Trials")
Publication.to_excel('Publication.xlsx', index=False, sheet_name="Publication")

# Partie : ajouter les auteurs

# Ouvrir le fichier Excel
wb = openpyxl.load_workbook('Publication.xlsx')
ws = wb.active

# Créer une variable partagée pour le verrouillage
lock = threading.Lock()

# Boucle à travers chaque ligne dans la colonne contenant les DOIs
rows = range(2, ws.max_row + 1)  # On commence à la ligne 2 car la ligne 1 contient les titres des colonnes
with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
    # Lancer chaque ligne dans un thread pour traiter plusieurs requêtes simultanément
    futures = {executor.submit(process_row, row): row for row in rows}

    # Parcourir les résultats des threads
    for future in concurrent.futures.as_completed(futures):
        # Récupérer le numéro de ligne associé au future en cours de traitement
        row = futures[future]
        try:
            result = future.result()
        except Exception as exc: # En cas d'exception, afficher un message d'erreur
            print(f"Processing row {row} generated an exception: {exc}")

# Enregistrer les modifications dans le fichier Excel
wb.save('Publication.xlsx')